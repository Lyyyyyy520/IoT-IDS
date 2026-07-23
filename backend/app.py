"""
IoT IDS Backend — Flask API Server v2.0
"""
from flask import Flask, jsonify, session, request
from flask_cors import CORS
from datetime import datetime
import os

app = Flask(__name__)
app.secret_key = os.urandom(24).hex()  # Session encryption key
app.config['SESSION_PERMANENT'] = True
CORS(app, supports_credentials=True)

# Init database on startup
from database import init_db, query_all, query_one, execute
init_db()

# Import auth services
from services.auth import require_auth, require_admin, login_user, logout_user, get_current_user, log_action

# Register probe blueprint
from api.probe import probe_bp
app.register_blueprint(probe_bp)

# ---- Health Check ----
_MODEL_PATH = os.path.join(os.path.dirname(__file__), 'data', 'best_model.onnx')
_model_loaded = os.path.exists(_MODEL_PATH)

@app.route('/api/health')
def health():
    return jsonify({
        'status': 'ok',
        'model_loaded': _model_loaded,
        'model_path': _MODEL_PATH if _model_loaded else None,
        'uptime': 0,
        'timestamp': datetime.now().isoformat(),
    })

# ---- Authentication Routes ----
@app.route('/api/auth/login', methods=['POST'])
def auth_login():
    data = request.get_json() or {}
    username = data.get('username', '').strip()
    password = data.get('password', '')
    if not username or not password:
        return jsonify({'success': False, 'message': '请输入账号和密码'}), 400
    result = login_user(username, password)
    if result['success']:
        return jsonify(result)
    return jsonify(result), 401

@app.route('/api/auth/logout', methods=['POST'])
def auth_logout():
    result = logout_user()
    return jsonify(result)

@app.route('/api/auth/me')
def auth_me():
    user = get_current_user()
    if not user:
        return jsonify({'authenticated': False}), 401
    return jsonify({'authenticated': True, 'user': user})


# ---- Dashboard Stats ----
def _get_traffic_history(source: str = ''):
    """统计最近 7 个 5 分钟间隔的流量"""
    result = []
    from datetime import datetime, timedelta
    src_filter = "AND source = ?" if source else ""
    src_params_extra = [source] if source else []

    for i in range(6, -1, -1):
        p1, p2 = f'-{(i+1)*5}', f'-{i*5}'
        total_row = query_one(
            f"SELECT COUNT(*) as total FROM traffic_logs "
            f"WHERE timestamp >= datetime('now', ? || ' minutes', 'localtime') "
            f"AND timestamp < datetime('now', ? || ' minutes', 'localtime') {src_filter}",
            [p1, p2] + src_params_extra,
        )
        atk_row = query_one(
            f"SELECT COUNT(*) as c FROM traffic_logs "
            f"WHERE timestamp >= datetime('now', ? || ' minutes', 'localtime') "
            f"AND timestamp < datetime('now', ? || ' minutes', 'localtime') "
            f"AND onnx_label IS NOT NULL AND onnx_label != '' AND onnx_label != 'normal' {src_filter}",
            [p1, p2] + src_params_extra,
        )
        total = total_row['total'] if total_row else 0
        atk = atk_row['c'] if atk_row else 0

        t = datetime.now() - timedelta(minutes=i*5)
        time_label = t.strftime('%H:%M')

        result.append({'time': time_label, 'normal': max(0, total - atk), 'attack': atk})

    return result


@app.route('/api/dashboard/stats')
def dashboard_stats():
    source = request.args.get('source', '')  # '' = all, 'sim' or 'real'
    src_prefix = f'[{source}]' if source else ''
    src_where = f"AND description LIKE '{src_prefix}%'" if source else ''
    tlog_where = f"WHERE source = '{source}'" if source else ''

    total_alerts = query_one(f"SELECT COUNT(*) as c FROM alerts WHERE 1=1 {src_where}")['c']
    alerts_today = query_one(
        f"SELECT COUNT(*) as c FROM alerts WHERE date(created_at) = date('now', 'localtime') {src_where}"
    )['c']
    active_threats = query_one(
        f"SELECT COUNT(DISTINCT src_ip) as c FROM alerts WHERE status = 'new' AND risk_level IN ('critical','high') {src_where}"
    )['c']
    total_assets = query_one("SELECT COUNT(*) as c FROM assets")['c']
    online_assets = query_one("SELECT COUNT(*) as c FROM assets WHERE status = 'online'")['c']
    total_traffic = query_one(f"SELECT COUNT(*) as c FROM traffic_logs {tlog_where}")['c'] if tlog_where else query_one("SELECT COUNT(*) as c FROM traffic_logs")

    dist_rows = query_all(
        f"SELECT onnx_label as label, COUNT(*) as count FROM traffic_logs "
        f"{tlog_where} AND onnx_label IS NOT NULL AND onnx_label != '' "
        f"GROUP BY onnx_label ORDER BY count DESC LIMIT 7"
    ) if tlog_where else query_all(
        "SELECT onnx_label as label, COUNT(*) as count FROM traffic_logs "
        "WHERE onnx_label IS NOT NULL AND onnx_label != '' "
        "GROUP BY onnx_label ORDER BY count DESC LIMIT 7"
    )
    type_name_map = {
        'normal': '正常流量', 'mirai': 'Mirai', 'gafgyt': 'Gafgyt',
        'other': '其他攻击',
        'ddos': 'Mirai', 'dos': 'Gafgyt', 'recon': '其他攻击', 'theft': '其他攻击',
    }
    attack_distribution = [
        {'type': type_name_map.get(r['label'], r['label']).title(), 'count': r['count']}
        for r in dist_rows
    ]

    recent = query_all(
        f"SELECT id, risk_level, attack_type, src_ip, dst_ip, confidence, created_at as timestamp, merged_count, status, description FROM alerts WHERE 1=1 {src_where} ORDER BY created_at DESC LIMIT 5"
    )

    import math
    def _count(level):
        r = query_one(
            f"SELECT COUNT(*) as c, COALESCE(AVG(confidence), 0.8) as avg_conf "
            f"FROM alerts WHERE risk_level = ? AND status = 'new' {src_where}", (level,)
        )
        return r['c'], r['avg_conf']

    crit_c, crit_conf = _count('critical')
    high_c, high_conf = _count('high')
    med_c, med_conf = _count('medium')
    low_c, low_conf = _count('low')

    threat_score = (
        8 * math.log(1 + crit_c * crit_conf) +
        4 * math.log(1 + high_c * high_conf) +
        1.5 * math.log(1 + med_c * med_conf) +
        0.3 * math.log(1 + low_c * low_conf)
    )

    # Source diversity penalty
    src_count = query_one("SELECT COUNT(DISTINCT src_ip) as c FROM alerts WHERE status = 'new'")['c']
    source_penalty = min(10, src_count * 1.0)

    # Defense bonus for blocked alerts
    total_all = query_one("SELECT COUNT(*) as c FROM alerts")['c']
    blocked = query_one("SELECT COUNT(*) as c FROM alerts WHERE status = 'blocked'")['c']
    defense_bonus = min(15, (blocked / max(1, total_all)) * 15) if total_all > 0 else 0

    risk_score = round(100 - threat_score - source_penalty + defense_bonus)
    risk_score = max(5, min(100, risk_score))

    return jsonify({
        'total_scanned': total_traffic,
        'alerts_today': alerts_today,
        'total_alerts': total_alerts,
        'active_threats': active_threats,
        'total_assets': total_assets,
        'online_assets': online_assets,
        'risk_score': risk_score,
        'system_status': 'normal' if risk_score > 60 else 'warning',
        'traffic_history': _get_traffic_history(source) if source else _get_traffic_history(),
        'attack_distribution': attack_distribution,
        'recent_alerts': recent,
    })


# ---- Alerts List ----
@app.route('/api/alerts')
def alerts_list():
    from flask import request
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 20, type=int)
    risk_filter = request.args.get('risk_level', '')
    type_filter = request.args.get('attack_type', '')
    time_filter = request.args.get('time_range', '')
    merged = request.args.get('merged', 'false').lower() == 'true'

    # 筛选条件
    where_parts = []
    where_params = []
    if risk_filter and risk_filter != 'all':
        where_parts.append("risk_level = ?")
        where_params.append(risk_filter)
    if type_filter and type_filter != 'all':
        where_parts.append("attack_type = ?")
        where_params.append(type_filter)
    if time_filter == '1h':
        where_parts.append("created_at >= datetime('now', '-1 hour', 'localtime')")
    elif time_filter == '24h':
        where_parts.append("created_at >= datetime('now', '-24 hour', 'localtime')")
    elif time_filter == '7d':
        where_parts.append("created_at >= datetime('now', '-7 days', 'localtime')")
    where_clause = ("WHERE " + " AND ".join(where_parts)) if where_parts else ""

    order_clause = """
        ORDER BY
            CASE risk_level WHEN 'critical' THEN 0 WHEN 'high' THEN 1 WHEN 'medium' THEN 2 ELSE 3 END,
            created_at DESC
    """

    if merged:
        # 合并模式：按 attack_type + src_ip + dst_ip 分组
        merge_sql = f"""
            SELECT
                MIN(a.id) as id,
                CASE
                    WHEN SUM(CASE WHEN a.risk_level = 'critical' THEN 1 ELSE 0 END) > 0 THEN 'critical'
                    WHEN SUM(CASE WHEN a.risk_level = 'high' THEN 1 ELSE 0 END) > 0 THEN 'high'
                    WHEN SUM(CASE WHEN a.risk_level = 'medium' THEN 1 ELSE 0 END) > 0 THEN 'medium'
                    ELSE 'low'
                END as risk_level,
                a.attack_type,
                a.src_ip,
                a.dst_ip,
                MAX(a.confidence) as confidence,
                MIN(a.created_at) as timestamp,
                COUNT(*) as merged_count,
                CASE
                    WHEN SUM(CASE WHEN a.status = 'blocked' THEN 1 ELSE 0 END) > 0 THEN 'blocked'
                    WHEN SUM(CASE WHEN a.status = 'false_positive' THEN 1 ELSE 0 END) > 0 THEN 'false_positive'
                    WHEN SUM(CASE WHEN a.status = 'new' THEN 1 ELSE 0 END) > 0 THEN 'new'
                    WHEN SUM(CASE WHEN a.status = 'reviewed' THEN 1 ELSE 0 END) > 0 THEN 'reviewed'
                    WHEN SUM(CASE WHEN a.status = 'resolved' THEN 1 ELSE 0 END) > 0 THEN 'resolved'
                    ELSE 'new'
                END as status,
                (SELECT a2.description FROM alerts a2
                 WHERE a2.attack_type = a.attack_type AND a2.src_ip = a.src_ip AND a2.dst_ip = a.dst_ip
                 ORDER BY a2.created_at DESC LIMIT 1) as description
            FROM alerts a
            {where_clause}
            GROUP BY a.attack_type, a.src_ip, a.dst_ip
        """
        merge_order = """
            ORDER BY
                CASE risk_level WHEN 'critical' THEN 0 WHEN 'high' THEN 1 WHEN 'medium' THEN 2 ELSE 3 END,
                timestamp DESC
        """
        count_sql = f"SELECT COUNT(*) as c FROM ({merge_sql})"
        total = query_one(count_sql, where_params)['c']

        offset = (page - 1) * page_size
        items_sql = f"SELECT * FROM ({merge_sql}) {merge_order} LIMIT ? OFFSET ?"
        items = query_all(items_sql, where_params + [page_size, offset])
    else:
        # 不合并模式：原始逐条显示
        count_sql = f"SELECT COUNT(*) as c FROM alerts {where_clause}"
        total = query_one(count_sql, where_params)['c']

        offset = (page - 1) * page_size
        items = query_all(
            f"SELECT id, risk_level, attack_type, src_ip, dst_ip, confidence, src_port, dst_port, protocol, "
            f"created_at as timestamp, 1 as merged_count, status, description "
            f"FROM alerts {where_clause} {order_clause} LIMIT ? OFFSET ?",
            where_params + [page_size, offset],
        )

    return jsonify({
        'total': total,
        'page': page,
        'page_size': page_size,
        'items': items,
    })


@app.route('/api/alerts/new')
def alerts_new():
    """查询新告警，用于前端实时轮询"""
    since_id = request.args.get('since_id', 0, type=int)

    items = query_all(
        "SELECT id, risk_level, attack_type, src_ip, dst_ip, confidence, "
        "created_at as timestamp, status, description "
        "FROM alerts WHERE id > ? "
        "ORDER BY id DESC LIMIT 20",
        (since_id,),
    )

    max_id_row = query_one("SELECT MAX(id) as max_id FROM alerts")
    max_id = max_id_row['max_id'] if max_id_row else 0

    return jsonify({'items': items, 'max_id': max_id})


# ---- Analysis Data ----
@app.route('/api/analysis/topology')
def topology_data():
    # 从 assets 表查所有设备作为节点
    assets = query_all("SELECT id, name, device_type, status, ip_address FROM assets")
    nodes = []
    for row in assets:
        risk_map = {
            'online': 'normal',
            'offline': 'high',
            'alert': 'critical',
        }
        nodes.append({
            'id': str(row['id']),
            'label': row['name'],
            'type': row['device_type'] if row['device_type'] else 'unknown',
            'risk': risk_map.get(row['status'], 'normal'),
            'ip': row['ip_address'],
        })

    # 从 alerts 表推导设备之间的连接关系
    links = []
    try:
        links_sql = """
            SELECT DISTINCT
                a1.id as source_id,
                a2.id as target_id
            FROM alerts al
            JOIN assets a1 ON al.src_ip = a1.ip_address
            JOIN assets a2 ON al.dst_ip = a2.ip_address
            WHERE a1.id != a2.id
            LIMIT 100
        """
        conns = query_all(links_sql)
        links = [{'source': str(row['source_id']), 'target': str(row['target_id'])} for row in conns]
    except Exception:
        pass

    return jsonify({
        'nodes': nodes,
        'links': links,
    })


@app.route('/api/analysis/heatmap')
def heatmap_data():
    # 统计最近 7 天、24 小时的告警分布
    sql = """
        SELECT
            strftime('%w', created_at) as dow,
            strftime('%H', created_at) as hour,
            COUNT(*) as count
        FROM alerts
        WHERE created_at >= datetime('now', '-7 days')
        GROUP BY dow, hour
        ORDER BY dow, hour
    """
    rows = query_all(sql)

    data = []
    for row in rows:
        data.append([int(row['hour']), int(row['dow']), row['count']])

    days = ['周日', '周一', '周二', '周三', '周四', '周五', '周六']
    hours = [f'{h:02d}:00' for h in range(24)]

    return jsonify({
        'data': data,
        'days': days,
        'hours': hours,
    })


@app.route('/api/analysis/mitre')
def mitre_data():
    """根据告警数据动态生成 MITRE ATT&CK 链路"""
    # 攻击类型 → MITRE 阶段映射
    stage_map = {
        'Mirai': 'Execution',
        'Gafgyt': 'Execution',
        'Other': 'Exfil',
        'Normal': 'Recon',
    }
    # 统计各阶段告警数
    rows = query_all("SELECT attack_type FROM alerts")
    stage_counts = {}
    for r in rows:
        stage = stage_map.get(r['attack_type'], 'Exfil')
        stage_counts[stage] = stage_counts.get(stage, 0) + 1

    stages_def = [
        {'name': '初始侦查', 'mitre': 'Recon', 'desc': '端口扫描探测'},
        {'name': '初始访问', 'mitre': 'InitAccess', 'desc': '漏洞利用与暴力破解'},
        {'name': '漏洞利用', 'mitre': 'Execution', 'desc': '远程代码/命令执行'},
        {'name': 'C2通信', 'mitre': 'C2', 'desc': 'C2服务器通信与控制'},
        {'name': '数据窃取', 'mitre': 'Exfil', 'desc': '敏感数据外传'},
    ]

    max_count = max(stage_counts.values()) if stage_counts else 1
    stages = []
    for s in stages_def:
        cnt = stage_counts.get(s['mitre'], 0)
        stages.append({
            **s,
            'alert_count': cnt,
            'active': cnt > 0,
            'intensity': round(cnt / max_count * 100),
        })

    total_alerts = sum(stage_counts.values()) or 1
    links = []
    for i in range(len(stages_def) - 1):
        src = stages_def[i]['mitre']
        tgt = stages_def[i + 1]['mitre']
        links.append({
            'source': src,
            'target': tgt,
            'value': round((stage_counts.get(src, 0) + stage_counts.get(tgt, 0)) / total_alerts * 100),
        })

    return jsonify({'stages': stages, 'links': links})


# ---- Detection ----
@app.route('/api/detect/upload', methods=['POST'])
def detect_upload():
    from flask import request
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400

    file = request.files['file']
    filename = file.filename or 'unknown'

    # Save uploaded file
    import os, tempfile
    tmp_dir = tempfile.mkdtemp()
    filepath = os.path.join(tmp_dir, filename)
    file.save(filepath)

    # Run detection
    from services.feature_extract import FeatureExtractor
    from models.inference import get_engine

    extractor = FeatureExtractor()
    features_list = extractor.extract_from_pcap(filepath)
    engine = get_engine()

    results = []
    for feat in features_list:
        result = engine.predict(feat)
        results.append({
            'class_name': result['class_name'],
            'confidence': result['confidence'],
            'risk_level': result['risk_level'],
            'is_attack': result['is_attack'],
        })

    # Cleanup
    try:
        os.remove(filepath)
        os.rmdir(tmp_dir)
    except Exception:
        pass

    # Summarize
    attack_count = sum(1 for r in results if r['is_attack'])
    return jsonify({
        'job_id': os.urandom(8).hex(),
        'filename': filename,
        'total_samples': len(results),
        'attack_count': attack_count,
        'normal_count': len(results) - attack_count,
        'results': results,
    })


# ---- Export ----
@app.route('/api/export/excel')
def export_excel():
    from flask import Response, request
    from openpyxl import Workbook
    from io import BytesIO

    risk_filter = request.args.get('risk_level', '')
    type_filter = request.args.get('attack_type', '')
    source_filter = request.args.get('source', '')
    merged = request.args.get('merged', 'false').lower() == 'true'

    where = []
    params = []
    if risk_filter and risk_filter != 'all':
        where.append("risk_level = ?"); params.append(risk_filter)
    if type_filter and type_filter != 'all':
        where.append("attack_type = ?"); params.append(type_filter)
    if source_filter == 'sim':
        where.append("description LIKE '[仿真]%'")
    elif source_filter == 'real':
        where.append("description LIKE '[真实]%'")
    w = ("WHERE " + " AND ".join(where)) if where else ""

    if merged:
        sql = f"SELECT MIN(id) as id,risk_level,attack_type,src_ip,dst_ip,MAX(confidence) as confidence,MAX(created_at) as created_at,COUNT(*) as cnt,status,description FROM alerts {w} GROUP BY attack_type,src_ip,dst_ip ORDER BY MAX(created_at) DESC LIMIT 50000"
    else:
        sql = f"SELECT id,risk_level,attack_type,src_ip,dst_ip,confidence,created_at,status,description FROM alerts {w} ORDER BY created_at DESC LIMIT 50000"

    wb = Workbook()
    ws = wb.active
    ws.title = '告警记录'
    if merged:
        ws.append(['ID', '风险等级', '攻击类型', '源IP', '目标IP', '置信度', '重复次数', '时间', '状态', '描述'])
    else:
        ws.append(['ID', '风险等级', '攻击类型', '源IP', '目标IP', '置信度', '时间', '状态', '描述'])
    alerts = query_all(sql, params)
    for a in alerts:
        ws.append([a.get('id',''), a['risk_level'], a['attack_type'],
                    a['src_ip'], a['dst_ip'], a['confidence'], a['created_at'],
                    a.get('cnt', 1), a['status'], a['description']])

    from openpyxl.styles import Font, PatternFill
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='2B5797', end_color='2B5797', fill_type='solid')

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=iot_ids_alerts.xlsx'})


@app.route('/api/data/cleanup', methods=['POST'])
def data_cleanup():
    """清空全部历史数据"""
    from database import get_config
    data = request.get_json() or {}
    if data.get('all'):
        # 立即清空全部
        execute("DELETE FROM alerts")
        execute("DELETE FROM traffic_logs")
        execute("DELETE FROM audit_logs")
        return jsonify({'success': True, 'message': '已清空全部历史数据'})
    # 按保留天数清理
    days = int(get_config('retention_days', '30'))
    execute("DELETE FROM alerts WHERE created_at < datetime('now', ? || ' days', 'localtime')", (f'-{days}',))
    execute("DELETE FROM traffic_logs WHERE timestamp < datetime('now', ? || ' days', 'localtime')", (f'-{days}',))
    execute("DELETE FROM audit_logs WHERE created_at < datetime('now', ? || ' days', 'localtime')", (f'-{days}',))
    return jsonify({'success': True, 'message': f'已清理 {days} 天前的历史数据'})


# ---- Traffic Capture Control ----
from services.traffic_capture import get_capture

@app.route('/api/capture/start', methods=['POST'])
def capture_start():
    data = request.get_json() or {}
    use_scapy = data.get('use_scapy', False)
    attack_ratio = data.get('attack_ratio', 0.25)
    result = get_capture().start(use_scapy=use_scapy, attack_ratio=attack_ratio)
    log_action('capture_start', f'mode={result["mode"]}')
    return jsonify(result)


@app.route('/api/capture/stop', methods=['POST'])
def capture_stop():
    result = get_capture().stop()
    log_action('capture_stop', f'packets={result["packet_count"]}')
    return jsonify(result)


@app.route('/api/capture/status')
def capture_status():
    return jsonify(get_capture().status())


# ---- Traffic Logs ----
@app.route('/api/traffic/logs')
def traffic_logs():
    source = request.args.get('source', '')
    if source:
        rows = query_all(
            "SELECT id, timestamp, src_ip, dst_ip, src_port, dst_port, protocol, length, flags "
            "FROM traffic_logs WHERE source = ? ORDER BY timestamp DESC LIMIT 50",
            (source,)
        )
    else:
        rows = query_all(
            "SELECT id, timestamp, src_ip, dst_ip, src_port, dst_port, protocol, length, flags "
            "FROM traffic_logs ORDER BY timestamp DESC LIMIT 50"
        )
    return jsonify({'items': rows})


# ---- Alert Actions ----
@app.route('/api/alerts/<int:alert_id>/block', methods=['POST'])
def block_ip(alert_id):
    """拉黑 IP：写入 policies 黑名单 + 更新告警状态 + 记录审计日志"""
    alert = query_one(
        "SELECT id, src_ip, attack_type, dst_ip, description FROM alerts WHERE id = ?",
        (alert_id,),
    )
    if not alert:
        return jsonify({'success': False, 'message': '告警不存在'}), 404

    src_ip = alert['src_ip']
    attack_type = alert['attack_type']

    # 检查是否已在黑名单中
    existing = query_one(
        "SELECT id FROM policies WHERE policy_type = 'blacklist' AND target = ?",
        (src_ip,),
    )
    if not existing:
        execute(
            "INSERT INTO policies (policy_type, target, action, description, enabled) VALUES (?, ?, ?, ?, 1)",
            ('blacklist', src_ip, 'block', f'来自告警 #{alert_id}: {attack_type} 攻击，目标 {alert["dst_ip"]}'),
        )

    # 更新该 IP 的所有告警状态
    execute(
        "UPDATE alerts SET status = 'blocked', updated_at = datetime('now', 'localtime') WHERE src_ip = ?",
        (src_ip,),
    )

    # 审计日志
    log_action('block_ip', f'alert_id={alert_id}, ip={src_ip}, attack={attack_type}')

    return jsonify({'success': True, 'message': f'已拉黑 IP: {src_ip}'})


@app.route('/api/alerts/<int:alert_id>/unblock', methods=['POST'])
def unblock_ip(alert_id):
    """解除拉黑：删除黑名单 policy + 恢复该 IP 所有告警状态"""
    alert = query_one("SELECT id, src_ip FROM alerts WHERE id = ?", (alert_id,))
    if not alert:
        return jsonify({'success': False, 'message': '告警不存在'}), 404

    ip_address = alert['src_ip']

    # 检查该 IP 是否有被拉黑的告警
    blocked_count = query_one(
        "SELECT COUNT(*) as c FROM alerts WHERE src_ip = ? AND status = 'blocked'",
        (ip_address,),
    )['c']
    if blocked_count == 0:
        return jsonify({'success': False, 'message': '该 IP 没有被拉黑的告警'}), 400

    execute("DELETE FROM policies WHERE policy_type = 'blacklist' AND target = ?", (ip_address,))
    execute(
        "UPDATE alerts SET status = 'reviewed', updated_at = datetime('now', 'localtime') WHERE src_ip = ? AND status = 'blocked'",
        (ip_address,),
    )

    log_action('unblock_ip', f'alert_id={alert_id}, ip={ip_address}')

    return jsonify({'success': True, 'message': f'已解除对 {ip_address} 的拉黑'})


@app.route('/api/alerts/<int:alert_id>/trace', methods=['POST'])
def trace_alert(alert_id):
    """溯源分析：生成溯源报告 + 更新告警状态"""
    alert = query_one(
        "SELECT id, src_ip, dst_ip, attack_type, risk_level, confidence, description, created_at "
        "FROM alerts WHERE id = ?",
        (alert_id,),
    )
    if not alert:
        return jsonify({'success': False, 'message': '告警不存在'}), 404

    # 查询同源 IP 的历史告警数
    history = query_one(
        "SELECT COUNT(*) as c FROM alerts WHERE src_ip = ? AND id != ?",
        (alert['src_ip'], alert_id),
    )

    # 查询该 IP 关联的设备信息
    asset = query_one(
        "SELECT name, device_type, mac_address FROM assets WHERE ip_address = ?",
        (alert['src_ip'],),
    )

    # 生成溯源报告
    trace_parts = [
        f"=== 溯源分析报告 ===\n",
        f"告警编号: #{alert['id']}",
        f"攻击来源 IP: {alert['src_ip']}",
        f"攻击目标 IP: {alert['dst_ip']}",
        f"攻击类型: {alert['attack_type']}",
        f"风险等级: {alert['risk_level']}",
        f"置信度: {alert['confidence']*100:.1f}%",
        f"首次发现: {alert['created_at']}",
    ]

    if asset:
        trace_parts.append(f"\n--- 关联设备信息 ---")
        trace_parts.append(f"设备名称: {asset['name']}")
        trace_parts.append(f"设备类型: {asset['device_type']}")
        trace_parts.append(f"MAC 地址: {asset['mac_address'] or '未知'}")

    if history:
        trace_parts.append(f"\n--- 历史关联 ---")
        trace_parts.append(f"同源 IP 历史告警: {history['c']} 条")
        trace_parts.append(f"该 IP 活跃程度: {'频繁' if history['c'] > 10 else '偶发' if history['c'] > 3 else '首次'}")

    # 攻击描述
    trace_parts.append(f"\n--- 攻击描述 ---")
    trace_parts.append(alert['description'] or f'{alert["attack_type"]} 网络攻击行为')

    trace_info = '\n'.join(trace_parts)

    # 更新告警
    execute(
        "UPDATE alerts SET trace_info = ?, status = 'reviewed', updated_at = datetime('now', 'localtime') WHERE id = ?",
        (trace_info, alert_id),
    )

    log_action('trace_alert', f'alert_id={alert_id}, src_ip={alert["src_ip"]}')

    return jsonify({
        'success': True,
        'trace_info': trace_info,
    })


@app.route('/api/alerts/<int:alert_id>/false-positive', methods=['POST'])
def mark_false_positive(alert_id):
    """标记误报：更新告警状态 + 记录审计日志"""
    alert = query_one("SELECT id, src_ip, attack_type FROM alerts WHERE id = ?", (alert_id,))
    if not alert:
        return jsonify({'success': False, 'message': '告警不存在'}), 404

    execute(
        "UPDATE alerts SET status = 'false_positive', updated_at = datetime('now', 'localtime') WHERE id = ?",
        (alert_id,),
    )

    log_action('mark_fp', f'alert_id={alert_id}, ip={alert["src_ip"]}, attack={alert["attack_type"]}')

    return jsonify({'success': True, 'message': f'已将告警 #{alert_id} 标记为误报'})


@app.route('/api/alerts/<int:alert_id>/unmark-false-positive', methods=['POST'])
def unmark_false_positive(alert_id):
    """撤销误报标记：恢复告警状态为 reviewed"""
    alert = query_one("SELECT id, status FROM alerts WHERE id = ?", (alert_id,))
    if not alert:
        return jsonify({'success': False, 'message': '告警不存在'}), 404
    if alert['status'] != 'false_positive':
        return jsonify({'success': False, 'message': '该告警未被标记为误报'}), 400

    execute(
        "UPDATE alerts SET status = 'reviewed', updated_at = datetime('now', 'localtime') WHERE id = ?",
        (alert_id,),
    )

    log_action('unmark_fp', f'alert_id={alert_id}')

    return jsonify({'success': True, 'message': f'已撤销告警 #{alert_id} 的误报标记'})


# ---- Blacklist Management ----
@app.route('/api/blocklist', methods=['GET'])
def blocklist_list():
    """获取所有黑名单记录（关联告警信息）"""
    # 对每个 IP 取风险最高的一条告警作为展示信息
    rows = query_all(
        "SELECT p.id, p.target AS ip_address, p.description AS reason, "
        "p.created_at AS blocked_at, p.enabled, "
        "a.id AS alert_id, a.attack_type, a.risk_level, a.dst_ip, a.status AS alert_status "
        "FROM policies p "
        "LEFT JOIN alerts a ON a.id = ("
        "  SELECT a2.id FROM alerts a2"
        "  WHERE a2.src_ip = p.target"
        "  ORDER BY CASE a2.risk_level WHEN 'critical' THEN 0 WHEN 'high' THEN 1 WHEN 'medium' THEN 2 ELSE 3 END,"
        "    a2.created_at DESC LIMIT 1"
        ") "
        "WHERE p.policy_type = 'blacklist' "
        "ORDER BY p.created_at DESC"
    )

    # 对同一 IP 的多条记录去重
    seen_ips = set()
    deduped = []
    for r in rows:
        if r['ip_address'] not in seen_ips:
            seen_ips.add(r['ip_address'])
            deduped.append(r)

    return jsonify({'total': len(deduped), 'items': deduped})


@app.route('/api/blocklist/<int:policy_id>', methods=['DELETE'])
def blocklist_delete(policy_id):
    """解除拉黑：删除 policy 记录 + 恢复关联告警状态"""
    policy = query_one(
        "SELECT id, target FROM policies WHERE id = ? AND policy_type = 'blacklist'",
        (policy_id,),
    )
    if not policy:
        return jsonify({'success': False, 'message': '黑名单记录不存在'}), 404

    ip_address = policy['target']

    # 删除该 IP 的所有黑名单 policy
    execute("DELETE FROM policies WHERE policy_type = 'blacklist' AND target = ?", (ip_address,))

    # 恢复该 IP 关联的告警状态
    execute(
        "UPDATE alerts SET status = 'reviewed', updated_at = datetime('now', 'localtime') WHERE src_ip = ? AND status = 'blocked'",
        (ip_address,),
    )

    log_action('unblock_ip', f'policy_id={policy_id}, ip={ip_address}')

    return jsonify({'success': True, 'message': f'已解除对 {ip_address} 的拉黑'})


# ---- Policy Management ----
@app.route('/api/policies', methods=['GET'])
def policies_list():
    """获取策略列表"""
    policy_type = request.args.get('type', '')
    where = "WHERE 1=1"
    params = []
    if policy_type:
        where += " AND policy_type = ?"
        params.append(policy_type)
    rows = query_all(
        f"SELECT * FROM policies {where} ORDER BY created_at DESC",
        params,
    )
    return jsonify({'items': rows})


@app.route('/api/policies', methods=['POST'])
def policies_create():
    """新增策略"""
    data = request.get_json() or {}
    policy_type = data.get('policy_type', 'blacklist')
    target = data.get('target', '')
    action = data.get('action', 'alert')
    description = data.get('description', '')
    enabled = data.get('enabled', 1)
    if not target:
        return jsonify({'success': False, 'message': '目标不能为空'}), 400
    pid = execute(
        "INSERT INTO policies (policy_type, target, action, description, enabled) VALUES (?,?,?,?,?)",
        (policy_type, target, action, description, enabled),
    )
    log_action('create_policy', f'id={pid}, type={policy_type}, target={target}')
    return jsonify({'success': True, 'id': pid})


@app.route('/api/policies/<int:policy_id>', methods=['PUT'])
def policies_update(policy_id):
    """编辑策略"""
    data = request.get_json() or {}
    execute(
        "UPDATE policies SET policy_type=?, target=?, action=?, description=?, enabled=? WHERE id=?",
        (data.get('policy_type'), data.get('target'), data.get('action'),
         data.get('description', ''), data.get('enabled', 1), policy_id),
    )
    log_action('update_policy', f'id={policy_id}')
    return jsonify({'success': True})


@app.route('/api/policies/<int:policy_id>', methods=['DELETE'])
def policies_delete(policy_id):
    """删除策略"""
    execute("DELETE FROM policies WHERE id = ?", (policy_id,))
    log_action('delete_policy', f'id={policy_id}')
    return jsonify({'success': True})


# ---- Asset Management ----
@app.route('/api/assets', methods=['GET'])
def assets_list():
    """获取设备列表"""
    rows = query_all("SELECT * FROM assets ORDER BY created_at DESC")
    return jsonify({'items': rows})


@app.route('/api/assets', methods=['POST'])
def assets_create():
    """新增设备"""
    data = request.get_json() or {}
    name = data.get('name', '').strip()
    ip = data.get('ip_address', '').strip()
    if not name or not ip:
        return jsonify({'success': False, 'message': '设备名和IP不能为空'}), 400
    aid = execute(
        "INSERT INTO assets (name, ip_address, mac_address, device_type, status, risk_level, last_seen) VALUES (?,?,?,?,?,?,datetime('now','localtime'))",
        (name, ip, data.get('mac_address', ''), data.get('device_type', 'other'),
         data.get('status', 'online'), data.get('risk_level', 'low')),
    )
    return jsonify({'success': True, 'id': aid})


@app.route('/api/assets/<int:asset_id>', methods=['PUT'])
def assets_update(asset_id):
    """编辑设备"""
    data = request.get_json() or {}
    execute(
        "UPDATE assets SET name=?, ip_address=?, mac_address=?, device_type=?, status=?, risk_level=?, last_seen=datetime('now','localtime') WHERE id=?",
        (data.get('name'), data.get('ip_address'), data.get('mac_address', ''),
         data.get('device_type'), data.get('status'), data.get('risk_level'), asset_id),
    )
    return jsonify({'success': True})


@app.route('/api/assets/<int:asset_id>', methods=['DELETE'])
def assets_delete(asset_id):
    """删除设备"""
    execute("DELETE FROM assets WHERE id = ?", (asset_id,))
    return jsonify({'success': True})


# ---- Log Center ----
@app.route('/api/logs/audit')
def logs_audit():
    """审计日志"""
    rows = query_all(
        "SELECT * FROM audit_logs ORDER BY created_at DESC LIMIT 200"
    )
    return jsonify({'items': rows})


@app.route('/api/logs/traffic')
def logs_traffic():
    """流量日志"""
    rows = query_all(
        "SELECT * FROM traffic_logs ORDER BY timestamp DESC LIMIT 200"
    )
    return jsonify({'items': rows})


# ---- Configuration ----
@app.route('/api/config', methods=['GET'])
def get_config():
    from database import get_config as gc
    return jsonify({
        'detection_mode': gc('detection_mode', 'offline'),
        'confidence_threshold': float(gc('confidence_threshold', '0.85')),
        'merge_window_minutes': int(gc('merge_window_minutes', '5')),
        'auto_block': gc('auto_block', 'false').lower() == 'true',
        'retention_days': int(gc('retention_days', '30')),
        'model_name': 'CNN+LSTM Hybrid',
        'model_version': 'v3.0',
        'input_features': 20,
        'output_classes': 4,
    })


@app.route('/api/config', methods=['PUT'])
def update_config():
    from database import set_config as sc
    data = request.get_json() or {}
    for key in ('detection_mode', 'confidence_threshold', 'merge_window_minutes', 'auto_block', 'retention_days'):
        if key in data:
            sc(key, str(data[key]).lower())
    return jsonify({'success': True})


if __name__ == '__main__':
    print('IoT IDS Backend starting...')
    print('    http://localhost:5000/api/health')
    app.run(host='0.0.0.0', port=5000, debug=True)
